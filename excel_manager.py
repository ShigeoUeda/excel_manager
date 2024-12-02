"""
シンプルなExcel操作ライブラリ

基本的なExcelファイルの読み書きと更新機能を提供します。
キー列を指定したデータの更新や、基本的な書式設定に対応しています。

Requirements:
    - Python 3.10+
    - openpyxl

Typical usage example:
    >>> excel = ExcelManager("data.xlsx")
    >>> excel.create_sheet("Sheet1", ["ID", "Name", "Value"])
    >>> excel.write_cell("Sheet1", 2, 1, "Data")
    >>> excel.save()
"""

# 標準ライブラリ
import os
import sys
import argparse
from pathlib import Path
from typing import List, Any, Optional, Tuple, Union

# サードパーティライブラリ
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

class ExcelManager:
    """
    Excelファイルを操作するためのシンプルなマネージャークラス
    
    このクラスは、Excelファイルの基本的な読み書き操作を提供します。
    新規ファイルの作成、既存ファイルの読み込み、データの書き込みと読み込み、
    セルの書式設定などの機能を備えています。

    Attributes:
        filename (str): 操作対象のExcelファイルパス
        wb (Workbook): openpyxlのWorkbookオブジェクト

    Raises:
        Exception: ファイルの読み込みに失敗した場合
    """

    def __init__(self, filename: str | Path) -> None:
        """
        ExcelManagerを初期化します。

        既存のファイルが存在する場合はそれを読み込み、
        存在しない場合は新規ファイルを作成します。

        Args:
            filename (str | Path): Excelファイルのパス

        Raises:
            Exception: ファイルの読み込みに失敗した場合
        """
        self.filename: str = str(filename)
        if os.path.exists(self.filename):
            try:
                self.wb: Workbook = load_workbook(self.filename)
                print(f"既存のファイル '{self.filename}' を読み込みました")
            except Exception as e:
                raise Exception(f"ファイル読み込みエラー: {str(e)}")
        else:
            self.wb = Workbook()
            print(f"新規ファイル '{self.filename}' を作成しました")

    def create_sheet(self, sheet_name: str, headers: Optional[List[str]] = None) -> Worksheet:
        """
        新規シートを作成し、必要に応じてヘッダーを設定します。

        ヘッダーが指定された場合、自動的に以下の書式が適用されます：
        - 太字
        - 中央揃え
        - グレーの背景色
        - 罫線
        - 列幅の自動調整

        Args:
            sheet_name (str): 作成するシート名
            headers (Optional[List[str]]): ヘッダー行のリスト

        Returns:
            Worksheet: 作成されたワークシート

        Raises:
            ValueError: 指定されたシート名が既に存在する場合
        """
        if sheet_name in self.wb.sheetnames:
            raise ValueError(f"シート '{sheet_name}' は既に存在します")
        
        ws: Worksheet = self.wb.create_sheet(title=sheet_name)
        if headers:
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                ws.column_dimensions[get_column_letter(col)].width = 15
        return ws

    def write_data(self, sheet_name: str, data: List[List[Any]], start_row: int = 1) -> None:
        """
        データをシートに書き込みます。

        2次元リストの形式でデータを受け取り、指定された行から順に書き込みます。
        すべてのセルに自動的に罫線が適用されます。

        Args:
            sheet_name (str): 書き込み先のシート名
            data (List[List[Any]]): 書き込むデータ（2次元リスト）
            start_row (int): 書き込み開始行（デフォルト: 1）

        Raises:
            ValueError: シートが存在しない場合
        """
        if sheet_name not in self.wb.sheetnames:
            raise ValueError(f"シート '{sheet_name}' が存在しません")
        
        ws: Worksheet = self.wb[sheet_name]
        for row_idx, row_data in enumerate(data, start_row):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

    def write_cell(
        self,
        sheet_name: str,
        row: int,
        column: Union[int, str],
        value: Any,
        number_format: Optional[str] = None
    ) -> None:
        """
        指定したセルにデータを書き込みます。

        列は数値（1始まり）または文字（'A'など）で指定できます。
        必要に応じて数値書式を指定することができます。

        Args:
            sheet_name (str): シート名
            row (int): 行番号（1から開始）
            column (Union[int, str]): 列番号（1から開始）または列記号（'A'など）
            value (Any): 書き込む値
            number_format (Optional[str]): 数値書式（例: '#,##0', 'yyyy/mm/dd'）

        Raises:
            ValueError: シートが存在しない場合、または無効な列指定の場合
        """
        if sheet_name not in self.wb.sheetnames:
            raise ValueError(f"シート '{sheet_name}' が存在しません")
        
        ws: Worksheet = self.wb[sheet_name]
        
        try:
            col_idx: int = column_index_from_string(column) if isinstance(column, str) else column
            if col_idx < 1:
                raise ValueError(f"無効な列番号です: {col_idx}")
            if row < 1:
                raise ValueError(f"無効な行番号です: {row}")
        except ValueError as e:
            raise ValueError(f"列指定が無効です: {str(e)}")
        
        cell = ws.cell(row=row, column=col_idx)
        cell.value = value
        if number_format:
            try:
                cell.number_format = number_format
            except ValueError as e:
                raise ValueError(f"無効な数値書式です: {number_format} - {str(e)}")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def write_cell_a1(
        self,
        sheet_name: str,
        cell_reference: str,
        value: Any,
        number_format: Optional[str] = None
    ) -> None:
        """
        A1形式でセルを指定してデータを書き込みます。

        Args:
            sheet_name (str): シート名
            cell_reference (str): セル参照（例: 'A1', 'B2'）
            value (Any): 書き込む値
            number_format (Optional[str]): 数値書式（例: '#,##0', 'yyyy/mm/dd'）

        Raises:
            ValueError: シートが存在しない場合、または無効なセル参照の場合

        Examples:
            >>> excel.write_cell_a1("Sheet1", "A1", 100)
            >>> excel.write_cell_a1("Sheet1", "B1", "2024/01/01", "yyyy/mm/dd")
        """
        if sheet_name not in self.wb.sheetnames:
            raise ValueError(f"シート '{sheet_name}' が存在しません")
        
        ws: Worksheet = self.wb[sheet_name]

        try:
            cell = ws[cell_reference]
        except (ValueError, KeyError) as e:
            raise ValueError(f"無効なセル参照です: {cell_reference} - {str(e)}")
        
        cell.value = value
        if number_format:
            try:
                cell.number_format = number_format
            except ValueError as e:
                raise ValueError(f"無効な数値書式です: {number_format} - {str(e)}")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def read_cell(
        self,
        sheet_name: str,
        row: int,
        column: Union[int, str]
    ) -> Any:
        """
        指定したセルの値を読み込みます。

        列は数値（1始まり）または文字（'A'など）で指定できます。

        Args:
            sheet_name (str): シート名
            row (int): 行番号（1から開始）
            column (Union[int, str]): 列番号（1から開始）または列記号（'A'など）

        Returns:
            Any: セルの値

        Raises:
            ValueError: シートが存在しない場合、または無効な列/行指定の場合
        """
        if sheet_name not in self.wb.sheetnames:
            raise ValueError(f"シート '{sheet_name}' が存在しません")
            
        ws: Worksheet = self.wb[sheet_name]
        
        try:
            col_idx: int = column_index_from_string(column) if isinstance(column, str) else column
            if col_idx < 1:
                raise ValueError(f"無効な列番号です: {col_idx}")
            if row < 1:
                raise ValueError(f"無効な行番号です: {row}")
        except ValueError as e:
            raise ValueError(f"列指定が無効です: {str(e)}")
            
        return ws.cell(row=row, column=col_idx).value

    def read_cell_a1(self, sheet_name: str, cell_reference: str) -> Any:
        """
        A1形式で指定したセルの値を読み込みます。

        Args:
            sheet_name (str): シート名
            cell_reference (str): セル参照（例: 'A1', 'B2'）

        Returns:
            Any: セルの値

        Raises:
            ValueError: シートが存在しない場合、または無効なセル参照の場合

        Examples:
            >>> value = excel.read_cell_a1("Sheet1", "A1")
            >>> value = excel.read_cell_a1("Sheet1", "B2")
        """
        if sheet_name not in self.wb.sheetnames:
            raise ValueError(f"シート '{sheet_name}' が存在しません")
            
        ws: Worksheet = self.wb[sheet_name]
        return ws[cell_reference].value

    def read_range(
        self,
        sheet_name: str,
        start_row: int,
        start_column: Union[int, str],
        end_row: int,
        end_column: Union[int, str]
    ) -> List[List[Any]]:
        """
        指定した範囲のデータを読み込みます。

        開始位置と終了位置を指定して、その範囲内のデータを2次元リストとして取得します。
        列は数値（1始まり）または文字（'A'など）で指定できます。

        Args:
            sheet_name (str): シート名
            start_row (int): 開始行
            start_column (Union[int, str]): 開始列
            end_row (int): 終了行
            end_column (Union[int, str]): 終了列

        Returns:
            List[List[Any]]: 読み込んだデータ（2次元リスト）

        Raises:
            ValueError: シートが存在しない場合、または無効な範囲指定の場合
        """
        if sheet_name not in self.wb.sheetnames:
            raise ValueError(f"シート '{sheet_name}' が存在しません")
            
        ws: Worksheet = self.wb[sheet_name]
        
        try:
            if start_row < 1 or end_row < 1:
                raise ValueError("行番号は1以上である必要があります")
            if end_row < start_row:
                raise ValueError("終了行は開始行以上である必要があります")
            
            start_col: int = column_index_from_string(start_column) if isinstance(start_column, str) else start_column
            end_col: int = column_index_from_string(end_column) if isinstance(end_column, str) else end_column
            
            if start_col < 1 or end_col < 1:
                raise ValueError("列番号は1以上である必要があります")
            if end_col < start_col:
                raise ValueError("終了列は開始列以上である必要があります")
        except ValueError as e:
            raise ValueError(f"範囲指定が無効です: {str(e)}")
            
        data: List[List[Any]] = []
        for row in ws.iter_rows(
            min_row=start_row,
            max_row=end_row,
            min_col=start_col,
            max_col=end_col,
            values_only=True
        ):
            data.append(list(row))
            
        return data

    def save(self) -> None:
        """
        ワークブックを保存します。

        変更内容をファイルに保存します。
        ファイルが他のプログラムで開かれている場合や、
        書き込み権限がない場合はエラーが発生します。

        Raises:
            PermissionError: ファイルへの書き込み権限がない場合
            OSError: ファイルの保存に失敗した場合
        """
        try:
            self.wb.save(self.filename)
        except PermissionError:
            raise PermissionError(f"ファイル '{self.filename}' への書き込み権限がありません")
        except OSError as e:
            raise OSError(f"ファイルの保存に失敗しました: {str(e)}")

def example_usage(filename: str) -> None:
    """
    使用例を示す関数

    Args:
        filename (str): 操作対象のExcelファイルパス
    """
    excel = ExcelManager(filename)
    
    # シートの作成
    excel.create_sheet("データ", ["ID", "名前", "値"])
    
    # データの書き込み（異なる方法）
    excel.write_cell("データ", 2, 1, 1)  # 行列指定
    excel.write_cell("データ", 2, "B", "サンプル1")  # 列記号指定
    excel.write_cell("データ", 2, 3, 1000, "#,##0")  # 数値書式指定
    
    excel.write_cell_a1("データ", "A3", 2)  # A1形式指定
    excel.write_cell_a1("データ", "B3", "サンプル2")
    excel.write_cell_a1("データ", "C3", 2000, "#,##0")
    
    # データの読み込み
    value1 = excel.read_cell("データ", 2, 1)
    value2 = excel.read_cell_a1("データ", "B2")
    
    # 範囲データの読み込み
    range_data = excel.read_range("データ", 2, "A", 3, "C")
    print(f"読み込んだデータ: {range_data}")
    
    excel.save()
    
def main() -> None:
    """CLI用のメイン関数"""
    parser = argparse.ArgumentParser(description='Excelファイル操作プログラム')
    parser.add_argument('-f', '--file', help='Excelファイルのパス', required=True)
    args = parser.parse_args()
    try:
        example_usage(args.file)
    except Exception as e:
        print(f"エラー: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main()
